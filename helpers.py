import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Font
import re
from cs50 import get_string
import time
import random
import calendar
import os

busyFileMidlands = ''
busyFileKlinicare = ''
busyFileRentmeester = ''
duplicateID = 1

def get_month(month):
    """Return name of the month"""
    month = int(month) - 1
    monthList = ["JANUARIE", "FEBRUARIE", "MAART", "APRIL", "MEI", "JUNIE", "JULIE", "AUGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]
    return monthList[int(month)]

def midlands(filename):
    """Return name of processed file"""
    #Prevent Race Conditions
    global busyFileMidlands
    if busyFileMidlands == filename:
        return "Error.txt"
    else:
        busyFileMidlands = filename
    #Time before the operations start
    then = time.time()
    #Load Excel Workbook
    print("Processing Excel Workbook")
    print("Warning: When inputting a Excel Workbook, make sure that there is no empty rows after the final row, otherwise you will get problems")
    print("Warning: When opening the Excel Workbook, make sure Sheet 1 isn't hidden")
    print("Creating Data Structures...")
    book =  openpyxl.load_workbook(filename)
    sheet = book["Sheet1"]
    #Initiates a list structure for customer accounts
    customerAccountsList = ["ACS001", "ACU001", "ADD01", "AGR002", "ALL003", "ALM002", "AMA003", "AMP001", "AND006", "ANI001", "ARA001", "ARE001H", "ARM005", "ASH001", "ASH003", "ASS001", "AVA001", "AVE001", "AWE001", "ASS001", "AVE001H", "AMJ001", "BAD001H", "BAK002", "BAL002", "BAL003", "BAN001", "BAN003", "BAR002", "BAR006", "BEC001H", "BEC002", "BED001", "BEN003", "BER001H", "BER003", "BER007", "BET001", "BET002", "BIE001", "BIG001", "BLU002", "BLO005", "BON001", "BON002", "BOS001", "BOS006", "BOS007", "BOT002", "BOT002H", "BOT004", "BOU001H", "BRA001", "BRA005", "BRE001", "BRO004", "BRU001 ", "BRU001H", "BUD001", "BUF001", "BUF002", "BUR003", "BUT001", "BUY001", "BYE001", "BAI001", "BAL002H", "BEN002", "BER001", "BOB001", "CAB001", "CAD001", "CAG001", "CAL001", "CAM001", "CAM003", "CAM003H", "CAM004", "CAM005H", "CAM006H", "CAM007", "CAN001", "CAP003", "CAS001H", "CAS002H", "CHA003", "CHA005", "CHO001", "CIP001", "CJG001", "CLA001H", "CLA002", "CLA002H", "CLA003H", "CLA004", "CLA006", "CLA008", "CLE001", "CLI002", "CLI003", "CLO001H", "CLO002", "COA001", "COA002", "COE001H", "COE003", "COE004", "COE008", "CON003", "CON004", "COO001", "COR001", "COR002H", "COR003H", "CRA003", "CRO010", "CUS001", "CUR001", "CWL001", "DAL001", "DAM001", "DAN009", "DAP001", "DAR001", "DAR003", "DAS001H", "DAV001", "DAV001H", "DAV005", "DDD001", "DEJ004H", "DEK002", "DEL004", "DEN001", "DEN002", "DEV004", "DEW001", "DEW001H", "DHD001", "DIE005", "DIP001", "DKJ001", "DOO001", "DOO004", "DOO002", "DRG001", "DUG001", "DUM001", "DUM002", "DUN004", "DUP003", "DUP004", "DUR001", "DRM001", "DUP003H", "DUP004H", "EDK001", "EDK001", "EDW002", "EKS001H", "EGV001", "ELS002", "EMA001", "EMA001H", "EME001", "EMP001", "EMS001", "ENG001", "ENG001H", "ENG003", "ENV001H", "EQU002", "EQU006", "EQU008", "ERA001H", "EST003", "ESH001", "ESH002", "ESS001", "EST001", "ETE001", "EUS001", "EVE004", "EXC001", "EXP001H", "FAI001", "FAI002", "FAM001", "FAR001", "FAR002", "FAR003", "FER003H", "FET001", "FRE001", "FOR002", "FOR003", "FOU001H", "FOU003", "FOU008", "FOU009", "FOY001", "FUL001", "FYV001H", "FER001H", "FYV001H", "GAM001", "GAR001", "GAR003", "GBM001", "GEN001", "GEO001", "GIA001", "GIB002", "GIG001H", "GIL002", "GIS001", "GLE001", "GOE001", "GOL002", "GON001", "GOP001", "GRA005", "GRA007", "GRA009", "GRA010", "GRE001", "GRE010", "GRI001", "GRI003", "GRO001", "GRO003", "GAL001", "GER003", "HAP001H", "HAR001H", "HAR003", "HAT005", "HEE001", "HEI001", "HEN001", "HEN004", "HEN008", "HER001", "HER004", "HEY004", "HFM001", "HIR001", "HLO001", "HLU001", "HOB001H", "HOL002", "HOP001", "HOU002", "HOW003", "HUM001", "HUM002", "HYT001", "HER002", "IBH001", "IMB001", "IMP002", "INA001", "IND002", "INT002", "INT003", "INV001", "JAB001H", "JAC002H", "JAN002", "JEL001H", "JEP001", "JHD001", "JOE001", "JOH003", "JOH005", "JOK001", "JON001", "JOO001", "JOR001", "JOS001H", "JOZ002", "JHS001", "JAM001H", "KAK001", "KAL001", "KER001", "KER002", "KIR001", "KJK001", "KLI002", "KLI003", "KLO001", "KLO002", "KOK001", "KOK002", "KRA003", "KRO001", "KUR001", "KZN001", "KZN002", "KZN003", "KLE007", "LAD002", "LAN001H", "LAN003", "LAU001H", "LEA001H", "LEE001", "LEE001H", "LER001", "LER002", "LER002H", "LER003H", "LEV001", "LES001", "LID001", "LIL001", "LOG001H", "LOM001H", "LOU001", "LOW003", "LOW004", "LUB001", "LUB001H", "LVA001", "LYL001", "LYN002", "LIP001H", "MAC001", "MAD001", "MAL001", "MAL003", "MAL004", "MAL006", "MAL007", "MAN002", "MAN001", "MAL005", "MAP003", "MAR002", "MAR004", "MAR005", "MAR007", "MAR009", "MAR010", "MAR011", "MAS001", "MAS002", "MAU001", "MEA003", "MEL001", "MEN001", "MEN002", "MER004", "MIC002", "MIE001", "MIK001", "MIL001", "MOL002", "MOL003", "MON008", "MON005", "MON007", "MOO002", "MOO003H", "MOO006", "MOR005", "MOS001", "MOT001", "MOU002", "MPA001", "MPP001", "MTU002", "MUD001", "MUT001", "MCK002", "NAM001", "NAT002", "NCA001", "NCA002", "NEL002H", "NEL003", "NEW002", "NHL001", "NIE001", "NOB001", "NOR005", "NOR009", "NOR012", "NOT009", "NOU001", "OAK004", "OAK001H", "OAT002H", "ODE002H", "OLD002", "OLI003", "ONE001", "OOS006", "ORA001", "ORA003", "PAR001", "PAR002", "PAR006", "PAR007", "PAR011", "PAT001", "PAT001H", "PAU001", "PAW002", "PAX002", "PEN003", "PET009", "PET012", "PFT001", "PHI001", "PIC002", "PID001", "PIE001", "PIE001H", "PIE002H", "PIE003H", "PIE006", "PIT002", "PJD001H", "PLA001", "PLA002", "POK002H", "PRE003", "PRI001", "PRI002", "PRI002H", "PRI005", "PRI006", "PRO006", "RAD002", "RAL001H", "RAS001", "RAT001", "RAT002", "RED001", "RIC004", "RIE001", "RIE002H", "ROB001", "ROB001H", "ROE001H", "ROS007", "ROS008", "ROU001", "ROU003", "ROU006", "ROV001", "ROY002", "ROSO012", "RAV001H", "REY001H", "SAD001", "SAN005", "SAP001", "SAP001H", "SAS001", "SCH002H", "SCH003H", "SCH003H", "SCH004", "SCH005", "SCO004", "SEL001", "SEL002", "SEN001", "SER002", "SEV002", "SHA003", "SIL002", "SIL003", "SKA002H", "SMA001", "SMA002", "SMI001", "SMI007", "SMI012", "SOM001", "SOU001", "SOU004", "SOU005", "SPC003", "SPC009", "SPC010", "SPC013", "SPC014", "SPE002", "SPO001", "SPR003", "SPR005", "SPR006", "SRC001", "STA004", "STA006", "STE002H", "STE009", "STE014", "STO006", "STR001H", "STO007", "STR002H", "SUM001", "SUM002", "SYM002", "SWA002H", "SWA004H", "SWA005", "SWA009", "SYK001", "SYM001", "SAP002H", "SOM002", "TAR002", "TAY001", "TEP001", "TER002", "TER005", "THE001", "THE002H", "THE003", "THE007", "THE008", "THE009", "THE012", "THE013", "THE019", "THO002", "TIG001", "TIN002", "TON001", "TOP001", "TRE003", "TRE004", "TRO001", "TUG001", "TUR001", "TRI001H", "UND001", "UND002", "UZU001", "VAL005", "VAN004H", "VAN005", "VAN007H", "VAN014", "VAN014H", "VAN016H", "VAN021", "VAN021H", "VAN023", "VAN028", "VAN033", "VAN038", "VAN040", "VAN048", "VAN050", "VAN069", "VAN072", "VEN001", "VEN001H", "VER004", "VER005", "VET002", "VET006", "VET007", "VET008", "VIL004", "VIL005", "VIS001H", "VIZ001", "VLA001H", "VOL001", "VOL002", "VRE002", "VRY001", "VRY003", "VAN006H", "VAN011H", "VEN002H", "WAL001", "WAL002", "WAL005", "WAR001", "WAT0001H", "WAT006", "WEL004", "WEN001", "WER001H", "WES001H", "WES013", "WES014", "WES016", "WES017", "WHI001", "WHI003", "WHI005", "WIC002", "WIL004", "WIL009", "WIL013", "WIL016", "WIN001", "WIN002", "WIN005", "WIN006", "WIT001", "WIT003", "WOL001", "WOO002", "WOR002", "WYN001"]
    #List for rows to be deleted
    deleteList = ["CASH SALES MOOI RIVER", "CASH SALES GAUTENG", "ADVOCIN INJ 100ML (S4)", "ALPHATRAK 2 TEST STRIPS 50 CT_1", "ANTISEDAN 10ML (S4)", "BRONCHICINE 10 DOSE (SINGLE)", "CERENIA INJ 20ML (S4)", "CERENIA TABS 160MG 4'S (S4)", "CERENIA TABS 16MG 4'S (S4)", "CERENIA TABS 24MG 4'S (S4)", "CERENIA TABS 60MG 4'S (S4)", "CLAMOXYL 200MG TABLETS 100'S  (S4)", "CLAMOXYL RTU 100ML (S4)", "CONVENIA LARGE 10ML (S4)     FRIDGE", "CONVENIA SMALL 4ML (S4)     FRIDGE", "DEFENSOR 3 (1 X 10ML VIAL)", "DEPO-MEDROL 40MG 5ML (S4)", "DEXDOMITOR INJ 10ML (S5)", "DOMITOR 10ML  (S5)", "DOMOSEDAN 5ML  (S5) ", "DRAXXIN 100ML  (S4)  ", "DRAXXIN 50ML (S4)", "EQUIVAC EHV-1/4  10DOSE (1X10ML)", "CIDR G DEVICE 20'S  (SHEEP)    V/O", "EXCENEL 4G PWDR  (S4) PLUS WATER", "EXCENEL RTU 100ML (S4)",
    "FELOCELL-3  (25X1D)  (CVR)", "FELOCELL-4  (25X1D)", "LEUKOCELL 2 (25X1D) ", "LUTALYSE 100ML  (S4)", "LUTALYSE 30ML  (S4)", "NUTRADYL TABS (100'S)", "PREDEF 2X  100ML  (S4)", "REVOLUTION 12% DOGS (0.25MLX3)2.6-5KG (PURPLE)",
    "REVOLUTION 12% DOGS (0.5MLX3)5.1-10KG (BROWN)", "REVOLUTION 12% DOGS (1MLX3)10.1-20KG (RED)", "REVOLUTION 12% DOGS (2MLX3)20.1-40KG (TEAL)", "REVOLUTION 6% CATS (0.75MLX3) 2.6-7.5KG (BLUE)" ,
    "REVOLUTION 6% KITT/PUP.(0.25MLX3) 0-2.5KG (PINK)", "RIMADYL CHEWABLE 100MG 180'S (S3)", "RIMADYL CHEWABLE 100MG 60'S  (S3)", "RIMADYL CHEWABLE 25MG 180'S  (S3)",
    "RIMADYL CHEWABLE 75MG 180'S (S3)", "RIMADYL INJ 20ML  (S3)     FRIDGE LINE", "RIMADYL INJ AQUEOUS 50ML  (S3)", "SYNULOX RTU 100ML  (S4)", "SYNULOX TAB 250MG 100'S  (S4)" ,
    "SYNULOX TAB 50MG 100'S  (S4)", "TERRACORTRIL EYE OINT.5ML SINGLE BOXED (S4)", "TROCOXIL 20MG (S4)", "TROCOXIL 30MG (S4)", "TROCOXIL 75MG (S4)",
    "TROCOXIL 95MG (S4)", "VANGUARD PLUS 5 (25 DOSE) ", "VANGUARD PLUS 5/CV 25D", "VANGUARD PLUS 5/CV-L 25D", "VANGUARD PLUS 5/L 25D", "WITNESS FELV+FIV (10'S)",
    "WITNESS PARVO (5'S) ", "ZENIQUIN TAB 100MG 50'S  (S4)", "ZENIQUIN TAB 25MG 100'S  (S4)", "EQUIVAC EHV-1/4 10DOSE (1X10ML)*EXPIRY 21/09/2018*",
    "EXCEDE INJ. 100ML (S4)", "PNEUMABORT-K+1b  20ML (1X10D) +STICKERS  #", "RIMADYL CHEWABLE 75MG 60'S  (S3)", "ZENIQUIN TAB 50MG 100'S  (S4)"]

    #Creates a dictionary structure of all the data inside of the excel workbook that needs to be replaced
    replaceDict = {
        "BODYGARD POUR ON 1LT"                              : ("BODYGARD POUR ON DIP 1L ", "1842"),
        "BODYGARD POUR ON 200ML"                            : ("BODYGARD POUR ON DIP 200ML ", "1840"),
        "BODYGARD POUR ON 500ML"                            : ("BODYGARD POUR ON DIP 500ML ", "1841"),
        "BOVI-SHIELD FP4+L5 (10D)                  #"       : ("BOVISHIELD FP4 + L5 10 DOSE ", "30893301"),
        "BOVI-SHIELD FP4+L5 (50D)                   #"      : ("BOVISHIELD FP4 + L5 50 DOSE ", "30893302"),
        "BOVI-SHIELD GOLD (10D)"                            : ("BOVISHIELD GOLD 5 10 DOSE ", "30875001"),
        "BOVI-SHIELD GOLD (50D)"                            : ("BOVISHIELD GOLD 5 50 DOSE ", "30875002"),
        "CIDR B DEVICE 10'S (WITHOUT CIDIROL)   V/O  #  "   : ("CIDR B CATTLE INSERT          ", "9953505"),
        "CIDR APPLICATOR FOR CATTLE"                        : ("CIDR CATTLE APPLICATORS       ", "9953506"),
        "CIDR APPLICATOR FOR SHEEP"                         : ("CIDR SHEEP & GOAT APPLICATOR  ", "31255401"),
        "CYDECTIN ORAL 5LT "                                : ("CYDECTIN ORAL 5 L ", "810128"),
        "CYDECTIN POUR ON 5LT"                              : ("CYDECTIN POUR ON 5 L ", "810143"),
        "CYDECTIN POUR ON 5LT SHORT DATED EXPIRY SEPT 2018" : ("CYDECTIN POUR ON 5 L ", "810143"),
        "DECTOMAX INJ 20ML                                #": ("DECTOMAX INJ SOLN 20 ML ", "38006498"),
        "DECTOMAX INJ 200ML              "                  : ("DECTOMAX INJ SOLN 200ML ", "30664603"),
        "DECTOMAX INJ 500ML                               #": ("DECTOMAX INJ SOLN 500ML ", "30664602"),
        "DECTOMAX INJ 50ML            #"                    : ("DECTOMAX INJ SOLN 50ML ", "30664601"),
        "EQUEST PLUS TAPE 11.8G "                           : ("EQUEST PLUS TAPE 1X12.2GM ", "810188"),
        "FARROWSURE GOLD (50 DOSE)"                         : ("FARROWSURE GOLD B 50 DOSE", "30884505"),
        "FOSTERA PCV 50D"                                   : ("FOSTERA PCV 50 DOSE", "813542"),
        "GLANVAC 3 100ML 100D FRIDGE"                       : ("GLANVAC 3 100ml ", "30695218"),
        "GLANVAC 3 250ML 250D FRIDGE"                       : ("GLANVAC 3 250ml ", "30695220"),
        "IMPROVAC 100ML (50DOSE) (FRIDGE)"                  : ("IMPROVAC (100ML x 50 DOSE)                                    ", "31283102"),
        "IMPROVAC 125D"                                     : ("IMPROVAC 125d", "31283105"),
        "INFORCE 3 10D (FRIDGE) + APPLICATOR"               : ("INFORCE-3 10 DOSE ", "30875020"),
        "INFORCE 3 50D + APPLICATOR"                        : ("INFORCE-3 50 DOSE ", "30875025"),
        "ENVIRACOR J-5 250ML 50D"                           : ("J5 ENVIRACOR ", "30876015"),
        "LECTADE SACHET SINGLE"                             : ("LECTADE 12 SACHETS ", "5402"),
        "LECTADE BOX OF 12"                                 : ("LECTADE 12 SACHETS ", "5402"),
        "ONE SHOT ULTRA-7 (10 DOSE) "                       : ("ONE SHOT ULTRA 7 10 DOSE ", "30874301"),
        "ONE SHOT ULTRA-7 (50 DOSE) "                       : ("ONE SHOT ULTRA 7 50 DOSE ", "30874302"),
        "PARACIDE 500ML"                                    : ("PARACIDE 500ML ", "1810"),
        "RESPISURE 100ML   (50 DOSE) FRIDGE"                : ("RESPISURE 50 DOSE ", "30695002"),
        "RESPISURE-ONE  100ML (50 DOSE)"                    : ("RESPISURE ONE 50 DOSE", "30838401"),
        "SCOURGUARD 4KC (10DOSE) 20ML"                      : ("SCOURGUARD 4KC 10DS ", "301310701"),
        "STARTECT 5LT"                                      : ("STARTECT 5LT ", "1291"),
        "SUPONA  AEROSOL 385ML  "                           : ("SUPONA AEROSOL 385ML", "813343"),
        "TERRAMYCIN 100 100ML      #"                       : ("TERRAMYCIN 100 INJ.SOLN.100ML", "30277603"),
        "TERRAMYCIN 100 50ML    #"                          : ("TERRAMYCIN 100 INJ.SOLN.50ML", "30277606"),
        "TERRAMYCIN LA 20ML                              #" : ("TERRAMYCIN L.A. INJ SOLN 20ML", "30474113"),
        "TERRAMYCIN LA 100ML                              #": ("TERRAMYCIN L.A. INJSOLN 100ML", "30474001"),
        "TERRAMYCIN LA 500ML                              #": ("TERRAMYCIN L.A. INJSOLN 500ML", "30475001"),
        "TERRAMYCIN WOUND SPRAY 150ML                  #"   : ("TERRAMYCIN WOUND SPRAY (AEROSOL 4G 150ML)", "30238103"),
        "ULTRA CHOICE-7 (50DOSE) 100ML"                     : ("ULTRA CHOICE 7 50 DOSE", "30838102"),
        "VALBANTEL 2LT"                                     : ("VALBANTEL 2L ", "1302"),
        "VALBANTEL 500ML #"                                 : ("VALBANTEL 500ML ", "1300"),
        "VALBAZEN CATTLE 200ML   #"                         : ("VALBAZEN FOR CATTLE 200ML ", "1100"),
        "VALBAZEN CATTLE 500ML  #"                          : ("VALBAZEN FOR CATTLE 500ML ", "1101"),
        "VALBAZEN CATTLE 5LT"                               : ("VALBAZEN FOR CATTLE 5L ", "1104"),
        "VALBAZEN SHEEP 500ML              #"               : ("VALBAZEN FOR SHEEP 500ML ", "1201"),
        "VALBAZEN SHEEP 5LT"                                : ("VALBAZEN FOR SHEEP 5L ", "1204"),
        "VALBAZEN ULTRA 2LT"                                : ("VALBAZEN ULTRA 2L ", "1282"),
        "VALBAZEN ULTRA 500ML #"                            : ("VALBAZEN ULTRA 500ML ", "1280"),
        "VALBAZEN ULTRA 5LT"                                : ("VALBAZEN ULTRA 5L ", "1285"),
        "VIBRIN (50 DOSE) 100ML         #"                  : ("VIBRIN 100ML ", "2812"),
        "RANOX 5LT"                                         : ("RANOX 5L ", "1635"),
        "CATTLEMASTER 4 (25 DOSE)"                          : ("CATTLEMASTER 4 ", "2502"),
        "VALBANTEL 10LT"                                    : ("VALBANTEL 10L ", "1310"),
        "VALBAZEN SHEEP 2LT"                                : ("VALBAZEN FOR SHEEP 2L ", "1203"),
        "CYDECTIN 50ML"                                     : ("CYDECTIN INJ 50 ML ", "810110"),
        "TERRAMYCIN 100 250ML"                              : ("TERRAMYCIN 100 INJ SOLN 250ML", "30277602"),
        "LITTERGUARD LTC 100ML (50 DOSE)"                   : ("LITTERGUARD LT-C 50 DOSE ", "30696702")
    }
    #Delete unnecessary columns
    print("Deleting unnecessary columns...")
    for i in range(7):
            sheet.delete_cols(1)
    for i in range(2):
        sheet.delete_cols(3)
    sheet.delete_cols(4)
    for i in range(2):
        sheet.delete_cols(5)
    sheet.delete_cols(6)
    for i in range(4):
        sheet.delete_cols(7)
    sheet.delete_cols(8)
    #Delete column descriptions
    print("Deleting column descriptions...")
    sheet.delete_rows(1)
    #Insert necessary columns
    print("Inserting necessary columns...")
    for i in range(2):
        sheet.insert_cols(1)
    for i in range(3):
        sheet.insert_cols(5)
    for i in range(4):
        sheet.insert_cols(9)
    for i in range(5):
        sheet.insert_cols(14)
    #Assigns maximum row and column length to x and y respectively
    x = sheet.max_row
    y = sheet.max_column
    #For every cell inside of the worksheet
    print("Processing Every Cell in the worksheet...")
    for row in range(1,x+1):
        for col in range(1, y+1):
            cell=sheet.cell(row=x+1-row,column=col)
            #Sets the styles of the excel worksheet
            cell.font = Font(name='Arial', size=10)
            #Process Column A
            if cell.column == 'A':
                cell.value = 'T'
            #Process Column B
            if cell.column == 'B':
                cell.value = 'NORM'
            #Process Column C
            if cell.column == 'C':
                if str(cell.value) not in customerAccountsList:
                    sheet.delete_rows(x+1-row)
                    break
            #Process Column D
            if cell.column == 'D':
                cell.value = (re.sub(r"[^a-zA-Z0-9]+", ' ', str(cell.value))).upper()
                if cell.value in deleteList:
                    sheet.delete_rows(x+1-row)
                    break
            #Process Column E
            if cell.column == 'E':
                cell.value = sheet.cell(row=x+1-row,column=col+16).value
                cell.number_format = 'YYYY/MM/DD'
                cell.value = (str(cell.value))[0:10]
                year = (str(cell.value))[0:4]
                month = (str(cell.value))[5:7]
                day = (str(cell.value))[8:10]
                cell.value = str(year) + str(month) + str(day)
            #Process Column F
            if cell.column == 'F':
                cell.value = (sheet.cell(row=x+1-row,column=col+14).value)
                cell.number_format = '0.00'
            #Process Column G
            if cell.column == 'G':
                cell.value = (sheet.cell(row=x+1-row,column=col+12).value)
            #Process Column J
            if cell.column == 'J':
                cell.value = '0'
            #Process Column L
            if cell.column == 'L':
                cell.value = "PFIZ"
            #Process Column M
            if cell.column == 'M':
                if cell.value in deleteList:
                    sheet.delete_rows(x+1-row)
                    break
                if cell.value in replaceDict:
                    key = cell.value
                    cell.value = replaceDict[key][0]
                    sheet.cell(row=x+1-row,column=col+2).value = replaceDict[key][1]
            #Process Column P
            if cell.column == 'P':
                cell.value = "MVW"
            #Process Column Q
            if cell.column == 'Q':
                cell.value = "MVW"
            #Process Column R
            if cell.column == 'R':
                cell.value = "PKEY"
    #Delete old columns
    print("Deleting old columns...")
    for i in range(3):
        sheet.delete_cols(19)
    #Adjust column widths
    print("Adjusting column widths...")
    sheet.column_dimensions['A'].width = 1.29 + 0.54
    sheet.column_dimensions['B'].width = 5.86 + 0.54
    sheet.column_dimensions['C'].width = 8.43 + 0.54
    sheet.column_dimensions['D'].width = 52.86 + 0.54
    sheet.column_dimensions['E'].width = 8.29 + 0.54
    sheet.column_dimensions['F'].width = 7.86 + 0.54
    sheet.column_dimensions['G'].width = 3.29 + 0.54
    sheet.column_dimensions['H'].width = 10.14 + 0.54
    sheet.column_dimensions['I'].width = 1.29 + 0.54
    sheet.column_dimensions['J'].width = 1.29 + 0.54
    sheet.column_dimensions['K'].width = 1.29 + 0.54
    sheet.column_dimensions['L'].width = 4.14 + 0.54
    sheet.column_dimensions['M'].width = 49.29 + 0.54
    sheet.column_dimensions['N'].width = 1.29 + 0.54
    sheet.column_dimensions['O'].width = 9.29 + 0.54
    sheet.column_dimensions['P'].width = 5 + 0.54
    sheet.column_dimensions['Q'].width = 5 + 0.54
    sheet.column_dimensions['R'].width = 5.43 + 0.54
    #Saves the workbook
    print("Saving Workbook...")
    book.save("BEWERKTEFILE_MVW_" + get_month(month) + year + ".xlsx")
    now = time.time() #Time after it finished
    print("Finished Processing Excel Workbook")
    print("It took: ", now-then, " seconds to process the Excel Workbook")
    busyFileMidlands = ''
    return "BEWERKTEFILE_MVW_" + get_month(month) + year + ".xlsx"

def klinicare(filename):
    """Return name of processed file"""
    #Prevent Race Conditions
    global busyFileKlinicare
    if busyFileKlinicare == filename:
        return "Error.txt"
    else:
        busyFileKlinicare = filename
    #Time before the operations start
    then = time.time()
    #Load Excel Workbook
    print("Processing Excel Workbook")
    print("Warning: When inputting a Excel Workbook, make sure that there is no empty rows after the final row, otherwise you will get problems")
    print("Warning: When opening the Excel Workbook, make sure Sheet 1 isn't hidden")
    print("Creating Data Structures...")
    book =  openpyxl.load_workbook(filename)
    sheet = book["MergedFile"]
    #Insert necessary rows needed before processing
    print("Inserting necessary columns...")
    sheet.insert_cols(9)
    sheet.insert_cols(10)
    sheet.insert_cols(11)
    #Assigns maximum row and column length to x and y respectively
    x = sheet.max_row
    y = sheet.max_column
    #Initiates variable for saving excel worksheet
    day = ''
    #For every cell inside of the worksheet
    print("Processing Every Cell in the worksheet...")
    for row in range(1,x+1):
        for col in range(1, y+1):
            cell=sheet.cell(row=x+1-row,column=col)
            #Process Column C
            if cell.column == 3:
                cell.value = re.sub(r"[^a-zA-Z0-9]+", ' ', str(cell.value))
                #Fixes Cellphone numbers of length 9 to be of length 10
                if len(str(cell.value)) == 9 and str(cell.value)[0] == '6' and str(cell.value)[1] == '2':
                    cellNr = "0"
                    cell.value = cellNr + str(cell.value)
                #Delete incorrect data
                if len(str(cell.value)) == 13:
                    #Skip over 13 digit numbers because they are correct
                    pass
                elif len(str(cell.value)) == 9 and str(cell.value)[0] == '0' or str(cell.value)[0] == '1' or str(cell.value)[0] == '2' or str(cell.value)[0] == '3':
                    sheet.delete_rows(x+1-row)
                    break
                elif len(str(cell.value)) == 9 and str(cell.value)[0] == '5' and str(cell.value)[1] == '1':
                    sheet.delete_rows(x+1-row)
                    break
                elif len(str(cell.value)) == 10 and str(cell.value)[0] != '0':
                    sheet.delete_rows(x+1-row)
                    break
                if str(cell.value)[0] == '2' and str(cell.value)[1] == '7' or len(str(cell.value)) in (17, 18, 19, 20):
                    sheet.cell(row=x+1-row,column=col+18).value = "CARD"
                    #Process 79 card numbers
                    if str(cell.value)[8] == '7' and str(cell.value)[9] == '9':
                        cell.value = str(cell.value[8:16])
                    elif str(cell.value)[9] == '7' and str(cell.value)[10] == '9':
                        cell.value = str(cell.value[9:17])
                    elif str(cell.value)[10] == '7' and str(cell.value)[11] == '9':
                        cell.value = str(cell.value[10:18])
                    elif str(cell.value)[11] == '7' and str(cell.value)[12] == '9':
                        cell.value = str(cell.value[11:19])
                    elif str(cell.value) == "79173058":
                        cell.value = "537592532"
                    elif str(cell.value) == "79763855":
                        cell.value = "537759320"
                    #Process 533 card numbers
                    elif str(cell.value)[8] == '5' and str(cell.value)[9] == '3' and str(cell.value)[10] == '3':
                        cell.value = str(cell.value[8:17])
                    elif str(cell.value)[9] == '5' and str(cell.value)[10] == '3' and str(cell.value)[11] == '3':
                        cell.value = str(cell.value[9:18])
                    elif str(cell.value)[10] == '5' and str(cell.value)[11] == '3' and str(cell.value)[12] == '3':
                        cell.value = str(cell.value[10:19])
                    elif str(cell.value)[11] == '5' and str(cell.value)[12] == '3' and str(cell.value)[13] == '3':
                        cell.value = str(cell.value[11:20])
                elif (str(cell.value)[0] == '7' and str(cell.value)[1] == '9' and len(str(cell.value)) == 8) or (str(cell.value)[0] == '5' and str(cell.value)[1] == '3' and str(cell.value)[2] == '3' and len(str(cell.value)) == 9):
                    sheet.cell(row=x+1-row,column=col+18).value = "CARD"
                else:
                    sheet.cell(row=x+1-row,column=col+18).value = "PROF"
                #Checks if data is valid, otherwise delete data out of the sheet
                if len(str(cell.value)) not in (6, 8, 9, 10, 13) or str(cell.value)[0].isalpha():
                    sheet.delete_rows(x+1-row)
                    break
            #Process Column D
            if cell.column == 4:
                if str(cell.value) == "123":
                    cell.value = sheet.cell(row=x+1-row,column=col-1).value
                if str(cell.value)[0] == 'O' and str(cell.value)[1] == 'T' and str(cell.value)[2] == 'C':
                    cell.value = sheet.cell(row=x+1-row,column=col-1).value
                if not cell.value:
                    cell.value = sheet.cell(row=x+1-row,column=col-1).value
                cell.value = re.sub(r"[^a-zA-Z0-9]+", ' ', str(cell.value))
                #Process Column E
            if cell.column == 5:
                year = (str(cell.value))[0:4]
                month = (str(cell.value))[4:6]
                if day == '':
                    day = (str(cell.value))[6:8]
                elif (str(cell.value))[6:8] > day:
                    day = (str(cell.value))[6:8]
            #Process Column F
            if cell.column == 6:
                cell.number_format = '0.00'
                if cell.value >= 0 and cell.value <= 0.99:
                    sheet.delete_rows(x+1-row)
                    break
                if sheet.cell(row=x+1-row,column=col-5).value == 'R':
                    cell.value *= -1
                    sheet.cell(row=x+1-row,column=col-5).value = 'T'
                    sheet.cell(row=x+1-row,column=col-4).value = "NORM"
            #Process Column G
            if cell.column == 7:
                if cell.value == '0':
                    cell.value = '1'
                if cell.value == '0.5':
                    cell.value == '1'
            #Process Column H
            if cell.column == 8:
                cell.value = str(cell.value) + '-' + str(sheet.cell(row=x+1-row,column=col+6).value)
            #Process Column J
            if cell.column == 10:
                cell.value = '0'
            #Process Column K
            if cell.column == 11:
                if str(sheet.cell(row=x+1-row,column=col+3).value) == '1':
                    cell.value = "FSHP"
                elif str(sheet.cell(row=x+1-row,column=col+3).value) == '2':
                    cell.value = "DISP"
            #Process Column M
            if cell.column == 13:
                cell.value = str(sheet.cell(row=x+1-row,column=col+3).value)
                cell.value = re.sub(r"[^a-zA-Z0-9]+", ' ', str(cell.value))
            #Process Column R
            if cell.column == 18:
                #Format every cell in Column R to number with 0 decimal places
                cell.number_format = '0'
                if cell.value == 101 :
                    sheet.delete_rows(x+1-row)
                cell.number_format = '@'
    #Delete and insert necessary Columns to format the worksheet as needed
    print("Deleting columns...")
    sheet.delete_cols(14)
    sheet.delete_cols(16)
    sheet.delete_cols(18)
    sheet.delete_cols(14)
    sheet.delete_cols(14)
    print("inserting columns...")
    sheet.insert_cols(14)
    sheet.insert_cols(17)
    #Sets Styles of excel worksheet
    #Assigns maximum row and column length to x and y respectively
    x = sheet.max_row
    y = sheet.max_column
    #For every cell inside of the worksheet
    print("Processing Every Cell in the worksheet...")
    for row in range(1,x+1):
        for col in range(1, y+1):
            cell=sheet.cell(row=x+1-row,column=col)
            cell.font = Font(name='Arial', size=10)
    #Adjust column widths
    print("Adjusting column widths...")
    sheet.column_dimensions['A'].width = 1.29 + 0.54
    sheet.column_dimensions['B'].width = 5.86 + 0.54
    sheet.column_dimensions['C'].width = 13.43 + 0.54
    sheet.column_dimensions['D'].width = 33.57 + 0.54
    sheet.column_dimensions['E'].width = 8.29 + 0.54
    sheet.column_dimensions['F'].width = 6.57 + 0.54
    sheet.column_dimensions['G'].width = 4.29 + 0.54
    sheet.column_dimensions['H'].width = 10.57 + 0.54
    sheet.column_dimensions['I'].width = 1.29 + 0.54
    sheet.column_dimensions['J'].width = 1.29 + 0.54
    sheet.column_dimensions['K'].width = 4.71 + 0.54
    sheet.column_dimensions['L'].width = 4.43 + 0.54
    sheet.column_dimensions['M'].width = 43.86 + 0.54
    sheet.column_dimensions['N'].width = 1.29 + 0.54
    sheet.column_dimensions['O'].width = 13.43 + 0.54
    sheet.column_dimensions['P'].width = 3.29 + 0.54
    sheet.column_dimensions['Q'].width = 1.29 + 0.54
    sheet.column_dimensions['R'].width = 5.14 + 0.54
    #Saves the worksheet
    print("Saving Workbook...")
    book.save("BEWERKTEFILE_KLINICARE_" + day + get_month(month) + year + ".xlsx")
    now = time.time() #Time after it finished
    print("Finished Processing Excel Workbook")
    print("It took: ", now-then, " seconds to process the Excel Workbook")
    busyFileKlinicare = ''
    return "BEWERKTEFILE_KLINICARE_" + day + get_month(month) + year + ".xlsx"

def rentmeester(filename):
    """Return name of processed file"""
    #Prevent Race Conditions
    global busyFileRentmeester
    if busyFileRentmeester == filename:
        return "Error.txt"
    else:
        busyFileRentmeester = filename
    #Time before the operations start
    then = time.time()
    book =  openpyxl.load_workbook(filename)
    sheet = book["Sheet2"]
    #Delete column descriptions
    print("Deleting column descriptions...")
    sheet.delete_rows(1)
    #Insert necessary rows needed before processing
    print("inserting rows...")
    for i in range(2):
        sheet.insert_cols(1)
    for i in range(4):
        sheet.insert_cols(4)
    for i in range(10):
        sheet.insert_cols(9)
    #Assigns maximum row and column length to x and y respectively
    x = sheet.max_row
    y = sheet.max_column
    #For every cell inside of the worksheet
    print("Processing Every Cell in the worksheet...")
    for row in range(1,x+1):
        for col in range(1, y+1):
            cell=sheet.cell(row=x+1-row,column=col)
            #Sets the styles of the excel worksheet
            cell.font = Font(name='Arial', size=10)
            #Process Column A
            if cell.column == 'A':
                cell.value = 'T'
            #Process Column B
            if cell.column == 'B':
                cell.value = 'NORM'
            #Process Column D
            if cell.column == 'D':
                cell.value = (str(sheet.cell(row=x+1-row,column=col+22).value) + ' ' + str(sheet.cell(row=x+1-row,column=col+23).value)).upper()
            #Process Column E
            if cell.column == 'E':
                cell.value = str(sheet.cell(row=x+1-row,column=col+18).value)
                cell.number_format = 'YYYY/MM/DD'
                cell.value = (str(cell.value))[0:7]
                year = (str(cell.value))[0:4]
                month = (str(cell.value))[5:7]
                day = calendar.monthrange(int(year), int(month))[1]
                cell.value = str(year) + str(month) + str(day)
            #Process Column F
            if cell.column == 'F':
                cell.value = round((sheet.cell(row=x+1-row,column=col+16).value * -0.01), 2)
                cell.number_format = '0.00'
            #Process Column G
            if cell.column == 'G':
                cell.value = '1'
            #Process Column J
            if cell.column == 'J':
                cell.value = '0'
            #Process Column L
            if cell.column == 'L':
                cell.value = "RENT"
            #Process Column M
            if cell.column == 'M':
                cell.value = get_month(month) + " " + "PREMIUM"
            #Process Column R
            if cell.column == 'R':
                cell.value = "PKEY"
    for i in range(12):
        sheet.delete_cols(19)
    #Adjust column widths
    print("Adjusting column widths...")
    sheet.column_dimensions['A'].width = 1.29 + 0.54
    sheet.column_dimensions['B'].width = 5.86 + 0.54
    sheet.column_dimensions['C'].width = 18.71 + 0.54
    sheet.column_dimensions['D'].width = 24.14 + 0.54
    sheet.column_dimensions['E'].width = 8.29 + 0.54
    sheet.column_dimensions['F'].width = 5.86 + 0.54
    sheet.column_dimensions['G'].width = 1.29 + 0.54
    sheet.column_dimensions['H'].width = 17.43 + 0.54
    sheet.column_dimensions['I'].width = 1.29 + 0.54
    sheet.column_dimensions['J'].width = 1.29 + 0.54
    sheet.column_dimensions['K'].width = 1.29 + 0.54
    sheet.column_dimensions['L'].width = 5.14 + 0.54
    sheet.column_dimensions['M'].width = 21.57 + 0.54
    sheet.column_dimensions['N'].width = 1.29 + 0.54
    sheet.column_dimensions['O'].width = 1.29 + 0.54
    sheet.column_dimensions['P'].width = 1.29 + 0.54
    sheet.column_dimensions['Q'].width = 1.29 + 0.54
    sheet.column_dimensions['R'].width = 5.43 + 0.54
    #Saves the worksheet
    print("Saving Workbook...")
    book.save("BEWERKTEFILE_RENT_" + get_month(month) + year + ".xlsx")
    now = time.time() #Time after it finished
    print("Finished Processing Excel Workbook")
    print("It took: ", now-then, " seconds to process the Excel Workbook")
    busyFileRentmeester = ''
    return "BEWERKTEFILE_RENT_" + get_month(month) + year + ".xlsx"